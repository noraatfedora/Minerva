from flask import (Blueprint, flash, g, redirect, render_template,
                   request, session, url_for, Flask, make_response, send_file
                   )
from sqlalchemy.sql.functions import current_date
from werkzeug.exceptions import abort
from minerva.backend.routes.auth import login_required, admin_required
import usaddress
import re
from json import loads, dumps
from collections import OrderedDict
from minerva.backend.apis import google_maps_qr
from minerva.backend.apis.assign import getUsers, measure
from minerva.backend.apis.db import users, conn, items, routes
from sqlalchemy import and_, select, desc
import pandas as pd
from os import environ
from datetime import date
from barcode import Code128
from openpyxl import load_workbook, styles
from barcode.writer import ImageWriter
from minerva.backend.apis import assign, order_assignment
import io
import pdfkit
import base64
import qrcode
import datetime
from fuzzywuzzy import fuzz
from minerva.backend.apis.email import send_recieved_notification, send_bagged_notification
bp = Blueprint('view_all_users', __name__)

addressFormat = ['AddressNumber', 'StreetNamePreDirectional', 'StreetNamePreModifier', 'StreetNamePreType',
                 'StreetName', 'StreetNamePostType', 'StreetNamePostModifier', 'StreetNamePostDirectional']


class Duplicate:
    # Generates a color for these cards
    # by setting the hue to a hash of the user ID's
    def generate_hue(self):
        return hash(sum(self.userIds)) % 255

    # Score here should be the max score
    # out of any combo in the list
    def __init__(self, userIds, score):
        self.userIds = userIds
        self.score = score
        self.hue = self.generate_hue()

    def __eq__(self, other):
        return self.userIds == other.userIds

    def __hash__(self):
        return sum(self.userIds)

    def __str__(self):
        return ("Duplicate with score " + str(self.score) + " containing " + str(self.userIds))


@login_required
@admin_required
@bp.route('/all_users', methods=('GET', 'POST'))
def all_users():
    # itemsList = conn.execute(items.select(items.c.foodBankId==g.user.foodBankId)).fetchall()
    userList = getUserList()
    showingDuplicates = False
    print(create_master_spreadsheet())
    if request.method == "POST" and 'num-vehicles' in request.values.to_dict().keys():
        print(request.values.to_dict())
        if 'redirect' in request.values.to_dict().keys():
            return loadingScreen(num_vehicles=request.values.get('num-vehicles'))
        else:
            assign.createAllRoutes(foodBankId=g.user.id, num_vehicles=int(
                request.values.get('num-vehicles')))
            return redirect('/all_orders')
    if request.method == "GET" and "volunteer" in request.args.keys():
        volunteerId = int(request.args.get("volunteer"))
        orderId = int(request.args.get("order"))
        order_assignment.assign(orderId=orderId, volunteerId=volunteerId)
        return redirect("/all_users")
    elif 'find-duplicates' in request.args.keys():
        # TODO: this is a really really shitty algorithm
        # and can definitley be sped up (but I only need to
        # run it once a week so it should be OK)
        setDuplicates = set()  # set of sets of user ID's
        for firstUser in userList:
            userIdSet = {firstUser.id}
            scoreMax = 0
            for secondUser in userList:
                score = fuzz.ratio(firstUser.name, secondUser.name)
                print("Ratio: " + str(score))
                # TODO: make this use things other than name
                if score > 80:
                    userIdSet.add(secondUser.id)
                    scoreMax = max(score, scoreMax)
            if len(userIdSet) > 1:
                setDuplicates.add(Duplicate(userIdSet, scoreMax))
        # print("Set duplicates:" + str(next(iter(setDuplicates))))
        userList = []
        for duplicate in setDuplicates:
            for userId in duplicate.userIds:
                row = (conn.execute(users.select().where(
                    users.c.id == userId)).fetchone())
                d = dict(row.items())
                d['hue'] = duplicate.hue
                userList.append(d)
        showingDuplicates = True

    if request.method == "POST":
        key = next(request.form.keys())
        print("Key: " + str(key))
        if "delete" in key:
            userId = int(key[len('delete-'):])
            conn.execute(users.delete().where(users.c.id == userId))
        if "enable" in key:
            userId = int(key[len('enable-'):])
            conn.execute(users.update().where(
                users.c.id == userId).values(disabled=False))
        if "download-master-spreadsheet" in key:
            return create_master_spreadsheet()
        if "disable" in key:
            userId = int(key[len('disable-'):])
            conn.execute(users.update().where(
                users.c.id == userId).values(disabled=True))
            routesList = conn.execute(routes.select().where(
                routes.c.foodBankId == g.user.foodBankId)).fetchall()
            for route in routesList:
                content = loads(route.content)
                if userId in content:
                    print("UserID in content!")
                    content.remove(userId)
                    conn.execute(routes.update().where(
                        routes.c.id == route.id).values(content=dumps(content)))
                    break
        userList = getUserList()

    volunteers = getVolunteers()
    today = datetime.date.today()
    activeUsersCount = 0
    disabledUsersCount = 0
    for user in userList:
        if user['disabled']:
            disabledUsersCount += 1
        else:
            activeUsersCount += 1

    return render_template("view_all_orders.html", users=userList, showingDuplicates=showingDuplicates, volunteers=volunteers, activeUsersCount=activeUsersCount, disabledUsersCount=disabledUsersCount)


def getUserList():
    return conn.execute(users.select().order_by(desc(users.c.disabled)).where(and_(users.c.foodBankId == g.user.id, users.c.role == "RECIEVER"))).fetchall()


@bp.route('/loading', methods=(['GET', 'POST']))
def loadingScreen(num_vehicles=40):
    return render_template("loading.html", num_vehicles=40)


@login_required
@admin_required
@bp.route('/routes-spreadsheet/', methods=('GET', 'POST'))
def send_spreadsheet():
    print("Sending spreadsheet...")
    google_maps = request.args.get('map') == 'true'
    routesList = conn.execute(routes.select(
        routes.c.foodBankId == g.user.id).order_by(routes.c.length)).fetchall()
    outputColumns = ['Number', 'First Name', "Last Name",
                     "Address", "Apt", "City", "State", "Zip", "Phone", "Notes"]
    if google_maps:
        outputColumns.append("Google Maps")
    pdList = []
    routeCount = 0
    for route in routesList:
        usersList = getUsers(route.id, addOriginal=True, includeDepot=True, columns=[
                             users.c.id, users.c.name, users.c.email, users.c.formattedAddress, users.c.address2, users.c.cellPhone, users.c.instructions])
        for user in usersList:
            userCount = 1
            try:
                parsed = usaddress.tag(user['Full Address'])[0]
                user['City'] = parsed['PlaceName']
                user['State'] = 'WA'
                user['Zip'] = parsed['ZipCode']
                address = ""
                for attribute in addressFormat:
                    if attribute not in parsed.keys():
                        continue
                    if address == "":
                        address = parsed[attribute]
                    else:
                        address = address + " " + parsed[attribute]
            except:
                address = user['Full Address']
            user['Address'] = address

            names = user['Name'].split(" ", maxsplit=1)
            user['First Name'] = names[0]
            user['Last Name'] = names[1]
            if user['Last Name'] == "nan":
                user['Last Name'] = ''
            if 'Google Maps' in outputColumns:
                user['Google Maps'] = google_maps_qr.make_single_url(
                    user['Full Address'])
            user['Number'] = userCount
        google_maps_link = google_maps_qr.make_url(usersList)
        # remove food bank
        usersList.remove(usersList[0])
        usersList.remove(usersList[len(usersList)-1])
        row_num = 15
        create_blank_rows(row_num - len(usersList), usersList, outputColumns)
        if google_maps:
            footerContent = [['', 'Google maps link:', google_maps_link], [
                '', 'Assigned to: ', ''], ['', 'Date:', ''], ['', 'Route ' + str(routeCount)]]
        else:
            footerContent = [['', 'Assigned to: ', ''], [
                '', 'Date:', ''], ['', 'Route ' + str(routeCount)]]
        routeCount += 1
        create_footer_rows(footerContent, usersList, outputColumns)
        df = pd.DataFrame(usersList)
        pdList.append(df)
    fileName = environ['INSTANCE_PATH'] + 'routes-' + getDateString() + '.xlsx'
    writer = pd.ExcelWriter(fileName, engine="openpyxl")
    for index in range(0, len(pdList)):
        pdList[index].to_excel(
            writer, sheet_name="Route " + str(index), index=False, columns=outputColumns)
    writer.save()

    # set formatting
    workbook = load_workbook(fileName)
    writer = pd.ExcelWriter(fileName, engine="openpyxl")
    writer.book = workbook
    for ws in workbook.worksheets:
        print(ws.title)
        for col in ws.iter_cols():
            maxWidth = ''
            for cell in col:
                cell.font = styles.Font(name='Times New Roman', size=10)
                if len(str(cell.value)) > len(maxWidth) and cell.value is not None and 'oogle.com' not in cell.value:
                    maxWidth = str(cell.value)
            ws.column_dimensions[col[0].column_letter].width = len(maxWidth)
    writer.save()
    return send_file(fileName, as_attachment=True)


@login_required
@admin_required
@bp.route('/routes-overview/', methods=('GET', 'POST'))
def send_overview():
    routesList = conn.execute(routes.select(
        routes.c.foodBankId == g.user.id).order_by(routes.c.length)).fetchall()
    dictList = []
    count = 0
    for route in routesList:
        dictList.append({
            'Route Number': count,
            'Length': len(loads(route.content)) - 2,
            'Distance': route.length / 1000,
            'Date': ' '
        })
        count += 1
    fileName = environ['INSTANCE_PATH'] + \
        'routes-overview-' + getDateString() + '.xlsx'
    df = pd.DataFrame(dictList)
    df.to_excel(fileName, index=False, header=True)
    return send_file(fileName, as_attachment=True)


def getDateString():
    return date.today().strftime('%m.%d')


def create_blank_rows(num_rows, currentList, outputColumns):
    for y in range(0, num_rows):
        toAppend = {}
        for x in outputColumns:
            toAppend[x] = ''
        currentList.append(toAppend)

# Content is a 2D list


def create_footer_rows(content, currentList, outputColumns):
    for rowList in content:
        rowDict = {}
        for x in range(0, len(rowList)):
            rowDict[outputColumns[x]] = rowList[x]
        for x in range(len(rowList), len(outputColumns)):
            rowDict[outputColumns[x]] = ''
        currentList.append(rowDict)


@login_required
@admin_required
@bp.route('/master-spreadsheet', methods=('GET', 'POST'))
def create_master_spreadsheet():
    columns = ['name', 'email', 'formattedAddress',
               'address2', 'cellPhone', 'instructions']
    prettyNames = {'formattedAddress': 'Full Address',
                   'address2': 'Apt',
                   'name': 'Name',
                   'email': 'Email',
                   'cellPhone': 'Phone',
                   'instructions': 'Notes'}
    outputColumns = ['First Name', "Last Name", "Email",
                     "Address", "Apt", "City", "Zip", "Phone", "Notes"]
    enabledRpList = conn.execute(users.select(and_(
        users.c.role == "RECIEVER", users.c.foodBankId == g.user.id, users.c.disabled == False))).fetchall()
    enabled = generateUserDataFrame(enabledRpList, prettyNames)
    disabledRpList = conn.execute(users.select(and_(users.c.role == "RECIEVER", users.c.foodBankId ==
                                  g.user.id, users.c.disabled == True)).order_by(users.c.disabledDate)).fetchall()
    if len(disabledRpList) > 0:
        disabled = generateUserDataFrame(disabledRpList, prettyNames)
    else:
        disabled = pd.DataFrame(columns=outputColumns)
    writer = pd.ExcelWriter(
        environ['INSTANCE_PATH'] + 'client-master-list.xlsx')
    enabled.to_excel(writer, sheet_name="Master list",
                     columns=outputColumns, startrow=0, index=False, na_rep="")
    disabled.to_excel(writer, sheet_name="Disabled clients",
                      columns=outputColumns, startrow=0, index=False, na_rep="")
    writer.save()
    return send_file(environ['INSTANCE_PATH'] + 'client-master-list.xlsx', as_attachment=True)


@login_required
@admin_required
@bp.route('/doordash-spreadsheet', methods=('GET', 'POST'))
def create_doordash_spreadsheet():
    def get_proximity(lat, lon):
        fb = conn.execute(users.select(
            users.c.id == g.user.foodBankId)).fetchone()
        fbLat = fb.latitude
        fbLon = fb.longitude
        return measure(lat, lon, fbLat, fbLon)

    #columns = ['name', 'email', 'formattedAddress', 'address2', 'cellPhone', 'instructions']
    prettyNames = {'formattedAddress': 'Full Address',
                   'address2': 'Apt',
                   'name': 'Name',
                   'email': 'Email',
                   'cellPhone': 'Phone'}
    routesRpList = conn.execute(routes.select(
        routes.c.foodBankId == g.user.id).order_by(routes.c.length)).fetchall()
    removeIds = []
    for route in routesRpList:
        removeIds.extend(loads(route.content))

    enabledRpList = conn.execute(users.select(and_(
        users.c.role == "RECIEVER", users.c.foodBankId == g.user.id, users.c.disabled == False))).fetchall()
    dictList = []
    staticValues = {
        'Pickup Location ID*': 'ELOISE-01',
        'Pickup Location Name': 'Eloise\'s Cooking Pot',
        'Pickup Window Start*': '10:00AM',
        'Pickup Window End*': '1:00PM',
        'Number of Items*': 1,
        'Timezone*': 'US/Pacific',
        'Dropoff Instructions (250 character max)': 'Please leave items, take a picture and leave. Do not return anything'
    }
    outputColumns = ['Pickup Location ID*', 'Order ID*', 'Pickup Location ID*', 'Date of Delivery*', 'Pickup Window Start*', 'Pickup Window End*', 'Timezone*', 'Client First Name*', 'Client Last Name*', 'Client Street Address*',
                     'Client Unit', 'Client City*', 'Client State*', 'Client ZIP*', 'Client Phone*', 'Number of Items*', 'Dropoff Instructions (250 character max)', 'Pickup Location Name', 'Pickup Phone Number', 'Pickup Instructions', 'Order Volume']

    for row in enabledRpList:
        if row.id in removeIds:
            continue
        columns = ['id', 'name', 'formattedAddress', 'address2',
                   'cellPhone', 'latitude', 'longitude']
        #columns = outputColumns

        def row2dict(r): return {c: betterStr(getattr(r, c)) for c in columns}
        d = row2dict(row)
        print(d)
        # doordash drivers can't go to the joint base
        if 'mcchord' not in d['formattedAddress'].lower():
            d['proxmity'] = get_proximity(d['latitude'], d['longitude'])
            try:
                d['Client First Name*'], d['Client Last Name*'] = d['name'].split(
                    ' ')
            except:
                d['Client First Name*'] = d['name']
                d['Client Last Name*'] = ''

            try:
                parsed = usaddress.tag(d['formattedAddress'])[0]
            except usaddress.RepeatedLabelError as e:
                print(e)
            d['Client City*'] = parsed['PlaceName']
            if 'ZipCode' in parsed:
                d['Client ZIP*'] = parsed['ZipCode']
            else:
                d['Client ZIP*'] = ''
            try:
                d['Client State*'] = parsed['StateName']
            except KeyError:
                d['Client State*'] = 'WA'
            address = ""
            for attribute in addressFormat:
                if attribute not in parsed.keys():
                    continue
                if address == "":
                    address = parsed[attribute]
                else:
                    address = address + " " + parsed[attribute]
            d['Client Street Address*'] = address
            d['Client Unit*'] = d.pop('address2')
            phone = d.pop('cellPhone')
            # Strip all non-numeric characters
            phone = re.sub('[^0-9]', '', phone)
            # Get last 10 digits
            phone = phone[-10:]
            if phone == "":
                phone = environ['DEFAULT_PHONE_NUM']
            d['Client Phone*'] = phone

            # get date as string seperated by '-'
            date = datetime.datetime.now().strftime("%Y-%m-%d")
            orderId = str(row.id) + "-" + date
            d['Order ID*'] = orderId
            d.update(staticValues)
            dictList.append(d)

    # Sort the dictList by proximity
    sortedList = sorted(dictList, key=lambda k: k['proxmity'], reverse=True)
    wednesdayList = []
    thursdayList = []
    # Sets wednesday to be the date of the next wednesday
    currentDate = datetime.datetime.today()
    wednesdayDate = currentDate + \
        datetime.timedelta((9-currentDate.weekday()) % 7)
    thursdayDate = currentDate + \
        datetime.timedelta((10-currentDate.weekday()) % 7)
    dateFormat = '%m/%d/%Y'
    for x in range(0, len(sortedList)):
        if x % 2 == 0:
            sortedList[x]['Date of Delivery*'] = wednesdayDate.strftime(
                dateFormat)
            wednesdayList.append(sortedList[x])
        else:
            sortedList[x]['Date of Delivery*'] = thursdayDate.strftime(
                dateFormat)
            thursdayList.append(sortedList[x])
    # Add empty row to wednesday list for spacing
    emptyDict = {}
    for key in outputColumns:
        emptyDict[key] = ''
    wednesdayList.append(emptyDict)
    outputDf = pd.DataFrame(wednesdayList + thursdayList)
    print(outputDf.columns)

    filename = 'doordash-' + getDateString() + '.xlsx'
    writer = pd.ExcelWriter(environ['INSTANCE_PATH'] + filename)
    outputDf.to_excel(writer, sheet_name="Doordash clients",
                      columns=outputColumns, startrow=0, index=False, na_rep="")
    writer.save()
    return send_file(environ['INSTANCE_PATH'] + filename, as_attachment=True)


def generateUserDataFrame(userRpList, prettyNames):
    columns = ['name', 'email', 'formattedAddress',
               'address2', 'cellPhone', 'instructions']
    def row2dict(r): return {prettyNames[c]: betterStr(
        getattr(r, c)) for c in columns}
    userDictList = []
    disabled = userRpList[0].disabled
    disabledDate = userRpList[0].disabledDate
    firstDate = True
    for user in userRpList:
        if (user.disabledDate != disabledDate or firstDate) and disabled:
            disabledDate = user.disabledDate
            blankRow = {'First Name': '', 'Last Name': '', 'Email': '', 'Address': '',
                        'Apt': '', 'City': '', 'Zip': '', 'Phone': '', 'Notes': ''}
            if disabledDate:
                # Type changes depending on if you're using mysql or sqlite apparently
                if type(disabledDate) == str:
                    removalRow = {'First Name': 'Removal', 'Last Name': disabledDate, 'Email': '',
                                  'Address': '', 'Apt': '', 'City': '', 'Zip': '', 'Phone': '', 'Notes': ''}
                else:
                    removalRow = {'First Name': 'Removal', 'Last Name': disabledDate.strftime(
                        '%m-%d-%y'), 'Email': '', 'Address': '', 'Apt': '', 'City': '', 'Zip': '', 'Phone': '', 'Notes': ''}
            else:
                removalRow = {'First Name': 'Removal', 'Last Name': '', 'Email': '',
                              'Address': '', 'Apt': '', 'City': '', 'Zip': '', 'Phone': '', 'Notes': ''}

            if not firstDate:
                userDictList.append(blankRow)
            userDictList.append(removalRow)
            firstDate = False
        userDict = row2dict(user)
        try:
            parsed = usaddress.tag(userDict['Full Address'])[0]
            userDict['City'] = parsed['PlaceName']
            userDict['Zip'] = parsed['ZipCode']
            address = ""
            for attribute in addressFormat:
                if attribute not in parsed.keys():
                    continue
                if address == "":
                    address = parsed[attribute]
                else:
                    address = address + " " + parsed[attribute]
            apt = user.address2
            if not apt:
                aptFormat = ['OccupancyType', 'OccupancyIdentifier']
                for attribute in aptFormat:
                    if attribute not in parsed.keys():
                        continue
                    if apt == "":
                        apt = parsed[attribute]
                    else:
                        apt = apt + " " + parsed[attribute]
                userDict['Apt'] = apt
                print("Setting apt value to " + apt)
                conn.execute(users.update().where(
                    users.c.id == user.id).values(address2=apt))
        except:
            address = userDict['Full Address']
        userDict['Address'] = address
        names = userDict['Name'].split(" ", maxsplit=1)
        userDict['First Name'] = names[0]
        if len(names) > 1:
            userDict['Last Name'] = names[1]
            if userDict['Last Name'] == "nan":
                userDict['Last Name'] = ''
        else:
            userDict['Last Name'] = ''
        userDictList.append(userDict)
    df = pd.DataFrame(userDictList)

    return df


def betterStr(value):
    if value == None:
        return ''
    else:
        return str(value)


@ login_required
@ admin_required
@ bp.route('/shipping_labels', methods=('GET', 'POST'))
def generate_shipping_labels():
    volunteers = getVolunteers()
    ordersDict = getOrders(g.user.id)
    itemsList = loads(conn.execute(users.select(
        users.c.id == g.user.foodBankId)).fetchone()['items'])
    html = render_template("shipping-labels.html", orders=ordersDict, volunteers=volunteers,
                           barcode_to_base64=barcode_to_base64, qrcode_to_base64=qrcode_to_base64)
    # Uncomment this line for debugging
    # return html
    pdf = pdfkit.from_string(html, False)
    response = make_response(pdf)
    response.headers['Content-type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline;'
    return response


def getVolunteers():
    proxy = conn.execute(users.select(and_(users.c.role == "VOLUNTEER",
                                           users.c.approved == True, users.c.volunteerRole == "DRIVER"))).fetchall()
    dictList = []
    for volunteer in proxy:
        volunteerDict = {}
        columns = conn.execute(users.select()).keys()
        for column in columns:
            volunteerDict[column] = getattr(volunteer, column)
        volunteerDict['numOrders'] = len(conn.execute(orders.select(
            and_(orders.c.volunteerId == volunteer.id, orders.c.completed == 0))).fetchall())
        dictList.append(volunteerDict)
    return dictList


def barcode_to_base64(orderId):
    imgByteArray = io.BytesIO()
    Code128(str(orderId) + "\n", writer=ImageWriter()).write(imgByteArray)
    return base64.b64encode(imgByteArray.getvalue()).decode()


def qrcode_to_base64(orderId):
    urlString = request.base_url[:-len('shipping_labels')] + \
        'auto_complete?orderId=' + str(orderId)
    imgByteArray = io.BytesIO()
    code = qrcode.make(urlString)
    code.save(imgByteArray, format="PNG")
    return base64.b64encode(imgByteArray.getvalue()).decode()
